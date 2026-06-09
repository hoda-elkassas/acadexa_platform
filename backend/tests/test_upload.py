"""
Unit and Integration Tests for Academic Record Upload & Import.
Tests Excel file parsing, route authorization, file validation, and import job processing.
"""
import os
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest
from fastapi import status
from fastapi.testclient import TestClient

from app.core.config import settings
from app.core.dependencies import get_current_user
from app.core.security import CurrentUser
from main import app
from app.services.file_parser import ExcelFileParser
from app.services.import_service import (
    ImportService,
    _process_student,
    calculate_semester_gpa,
    calculate_cumulative_gpa,
)

client = TestClient(app)


# ─── FIXTURES ──────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_excel_file():
    """Generates a valid temporary Excel file matching the academic record layout."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Sheet"

        # Student info headers
        ws["A1"] = "كود الطالب : 20230001"
        ws["A2"] = "أسم الطالب : محمد أحمد"
        ws["A3"] = "مستوى الدراسة : المستوى الأول"
        ws["A4"] = "النسبة(بحساب النقاط) : 85.0"

        # Semester block header (Row 40 offset base)
        # Column K (11): Section/Department
        # Column AH (34): Level/Semester
        # Column BO (67): Academic Year
        ws.cell(row=40, column=11, value="القسم/الشعبة : علوم الحاسب")
        ws.cell(row=40, column=34, value="المستوى/الفصل : المستوى الأول/الفصل الدراسي الأول")
        ws.cell(row=40, column=67, value="العام الأكاديمي   : 2022-2023")

        # Course 1 (Row 41)
        # Column CD (82): Index (digit)
        # Column BC (55): Course name
        # Column BU (73): Course code
        # Column I (9): Passed status (نعم)
        # Column O (15): Grade letter (A)
        # Column Q (17): Score (95)
        # Column AL (38): Credit hours (3)
        ws.cell(row=41, column=82, value="1")
        ws.cell(row=41, column=55, value="برمجة 1")
        ws.cell(row=41, column=73, value="CS101")
        ws.cell(row=41, column=9, value="نعم")
        ws.cell(row=41, column=15, value="A")
        ws.cell(row=41, column=17, value="95")
        ws.cell(row=41, column=38, value="3")

        wb.save(tmp.name)
        wb.close()
        tmp_path = tmp.name

    yield Path(tmp_path)

    # Cleanup after test
    if os.path.exists(tmp_path):
        os.remove(tmp_path)


# ─── PARSER TESTS ─────────────────────────────────────────────────────────────

def test_excel_parser_arabic_headers(sample_excel_file):
    """Verifies that ExcelFileParser correctly extracts student info and courses from the mock workbook."""
    parser = ExcelFileParser()
    data = parser.parse_workbook(sample_excel_file)

    assert not data.errors, f"Parser encountered unexpected errors: {data.errors}"
    assert len(data.students) == 1

    student = data.students[0]
    assert student.student_code == "20230001"
    assert student.name == "محمد أحمد"
    assert student.enrollment_year == 2023
    assert student.cumulative_percentage == 85.0
    assert len(student.semesters) == 1

    semester = student.semesters[0]
    assert semester.term == "fall"
    assert len(semester.courses) == 1

    course = semester.courses[0]
    assert course.course_code == "CS101"
    assert course.course_name == "برمجة 1"
    assert course.credit_hours == 3
    assert course.grade_letter == "A"
    assert course.score == 95.0
    assert course.passed is True


# ─── ROUTER / ENDPOINT TESTS ──────────────────────────────────────────────────

def test_upload_invalid_file_type():
    """Verifies that upload endpoint rejects file formats other than Excel with 400 Bad Request."""
    # Override authorization dependency mapping
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="user-123", email="user@example.com", system_role="DEVELOPER"
    )

    response = client.post(
        "/api/v1/upload/academic-record",
        data={"department_id": "dept-123"},
        files={"file": ("test.pdf", b"pdf content", "application/pdf")},
        headers={"Authorization": "Bearer mock-token"}
    )

    app.dependency_overrides.clear()
    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "Only Excel files (.xlsx, .xls) are allowed" in response.json()["detail"]


def test_upload_file_size_exceeded():
    """Verifies that files exceeding the MAX_UPLOAD_SIZE_MB trigger a 400 error."""
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        id="user-123", email="user@example.com", system_role="DEVELOPER"
    )

    # Temporarily set max upload size to 0 to force size check failure
    original_max = settings.MAX_UPLOAD_SIZE_MB
    settings.MAX_UPLOAD_SIZE_MB = 0

    try:
        response = client.post(
            "/api/v1/upload/academic-record",
            data={"department_id": "dept-123"},
            files={"file": ("test.xlsx", b"some bytes", "application/vnd.ms-excel")},
            headers={"Authorization": "Bearer mock-token"}
        )
    finally:
        settings.MAX_UPLOAD_SIZE_MB = original_max
        app.dependency_overrides.clear()

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert "File size exceeds the limit" in response.json()["detail"]


# ─── SERVICE / DATABASE TESTS ─────────────────────────────────────────────────

@patch("app.services.import_service.supabase_admin")
@patch("app.services.import_service.execute_rpc")
@patch("app.services.import_service.fetch_one")
@patch("app.services.import_service.fetch_many")
@pytest.mark.asyncio
async def test_import_service_student_creation(
    mock_fetch_many, mock_fetch_one, mock_execute_rpc, mock_supabase
):
    """Tests that _process_student runs the plan lookup, resolves grade scales, and updates students, semesters, and courses."""
    # 1. Mock RPC and Database fetches
    mock_execute_rpc.return_value = "plan-uuid"
    mock_fetch_one.side_effect = lambda table, filters: {
        "study_plans": {"id": "plan-uuid", "default_grading_scale_id": "scale-uuid"}
    }.get(table)

    mock_fetch_many.return_value = [
        {"grade_letter": "A", "points": 4.0},
        {"grade_letter": "B", "points": 3.0},
    ]

    # Mock DB table responses
    mock_students_table = MagicMock()
    mock_students_table.upsert.return_value.execute.return_value.data = [{"id": "student-uuid"}]

    mock_semesters_table = MagicMock()
    mock_semesters_table.upsert.return_value.execute.return_value.data = [
        {"semester_number": 1, "id": "sem-uuid"}
    ]

    mock_courses_table = MagicMock()
    mock_courses_table.select.return_value.eq.return_value.execute.return_value.data = []  # No existing courses
    mock_courses_table.select.return_value.execute.return_value.data = [{"id": "course-uuid", "code": "CS101"}]

    def side_effect(table_name):
        if table_name == "students":
            return mock_students_table
        elif table_name == "student_semesters":
            return mock_semesters_table
        elif table_name == "student_courses":
            return mock_courses_table
        elif table_name == "courses":
            return mock_courses_table
        return MagicMock()

    mock_supabase.table.side_effect = side_effect
    mock_supabase.rpc.return_value.execute.return_value.data = "student-course-uuid"

    # 2. Build mock ParsedStudent input
    from app.services.file_parser import ParsedStudent, ParsedSemester, ParsedCourse
    course = ParsedCourse(
        course_code="CS101",
        course_name="Introduction to Programming",
        credit_hours=3,
        grade_letter="A",
        score=95.0,
        passed=True,
        grade_points=4.0
    )
    semester = ParsedSemester(
        semester_number=1,
        academic_year="2022-2023",
        term="fall",
        level=1,
        department="علوم الحاسب",
        level_semester_raw="",
        courses=[course]
    )
    student = ParsedStudent(
        student_code="20230001",
        name="محمد أحمد",
        study_level_str="المستوى الأول",
        cumulative_percentage=85.0,
        enrollment_year=2023,
        department="علوم الحاسب",
        semesters=[semester],
        cumulative_gpa=4.0,
        total_passed_hours=3
    )

    success, err_msg = await _process_student(student, "dept-123", "job-123")

    assert success is True
    assert err_msg is None

    # Assert student and semesters table upserts were called
    mock_students_table.upsert.assert_called_once()
    mock_semesters_table.upsert.assert_called_once()


# ─── DUP / GPA CALCULATION TESTS ──────────────────────────────────────────────

def test_gpa_calculations():
    """Validates calculate_semester_gpa and calculate_cumulative_gpa formulas."""
    from app.services.file_parser import ParsedCourse, ParsedSemester
    scale = {"A": 4.0, "B": 3.0, "F": 0.0}

    courses = [
        ParsedCourse(course_code="C1", course_name="", credit_hours=3, grade_letter="A", passed=True, grade_points=4.0),
        ParsedCourse(course_code="C2", course_name="", credit_hours=4, grade_letter="B", passed=True, grade_points=3.0),
        ParsedCourse(course_code="C3", course_name="", credit_hours=3, grade_letter="W", passed=False, grade_points=0.0), # SPECIAL
    ]

    sem_gpa = calculate_semester_gpa(courses, scale)
    # total points: 4.0*3 + 3.0*4 = 24.0. total hours = 7. 24 / 7 = 3.4286
    assert sem_gpa == 3.4286

    semester1 = ParsedSemester(semester_number=1, academic_year="", term="", level=1, courses=courses)
    semester2 = ParsedSemester(
        semester_number=2, academic_year="", term="", level=1,
        courses=[
            ParsedCourse(course_code="C4", course_name="", credit_hours=3, grade_letter="A", passed=True, grade_points=4.0)
        ]
    )

    cum_gpa = calculate_cumulative_gpa([semester1, semester2], scale)
    # total points: 24.0 + 4.0*3 = 36.0. total hours = 10. 36 / 10 = 3.6
    assert cum_gpa == 3.6
