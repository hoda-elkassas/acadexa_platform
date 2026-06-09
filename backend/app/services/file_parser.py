"""
Excel File Parser for academic records.

Keeps the proven column-offset detection logic (detect_offset, get_cell_with_offset, COLS).
Adds: structured dataclasses, GPA calculation from grades, level/term parsing, retake detection.

RULE: This module does NO database access. It is a pure parser returning dataclasses.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger("acadexa.parser")

# ─── Constants ────────────────────────────────────────────────────────────────

SPECIAL_SYMBOLS = {"W", "FW", "IC", "MW", "AU", "S", "TC", "Ex", "IP", "NP", "I"}

FALLBACK_GRADE_POINTS: Dict[str, float] = {
    "A+": 4.0, "A": 4.0, "A-": 3.7,
    "B+": 3.4, "B": 3.2, "B-": 3.0,
    "C+": 2.8, "C": 2.5, "C-": 2.2,
    "D+": 1.8, "D": 1.5, "D-": 1.2,
    "F": 0.0,
}

LEVEL_MAP: Dict[str, int] = {
    "الاول": 1, "الأول": 1, "أول": 1, "اول": 1, "1": 1,
    "الثاني": 2, "الثانى": 2, "ثاني": 2, "ثانى": 2, "2": 2,
    "الثالث": 3, "ثالث": 3, "3": 3,
    "الرابع": 4, "رابع": 4, "4": 4,
}

TERM_MAP: Dict[str, str] = {
    "الفصل الدراسى الاول": "fall",
    "الفصل الدراسي الاول": "fall",
    "الفصل الدراسى الأول": "fall",
    "الفصل الدراسي الأول": "fall",
    "الفصل الأول": "fall",
    "الفصل الاول": "fall",
    "الفصل الدراسى الثاني": "spring",
    "الفصل الدراسي الثاني": "spring",
    "الفصل الدراسى الثانى": "spring",
    "الفصل الدراسي الثانى": "spring",
    "الفصل الثاني": "spring",
    "الفصل الثانى": "spring",
    "الفصل الصيفى": "summer",
    "الفصل الصيفي": "summer",
    "صيفي": "summer",
}



# ─── Dataclasses ──────────────────────────────────────────────────────────────

@dataclass
class ParsedCourse:
    course_code: str
    course_name: str
    credit_hours: int
    grade_letter: str
    score: Optional[float] = None
    passed: bool = False
    grade_points: float = 0.0
    is_retake: bool = False
    retake_count: int = 0


@dataclass
class ParsedSemester:
    semester_number: int
    academic_year: str = ""
    term: Optional[str] = None         # fall | spring | summer
    level: Optional[int] = None
    department: str = ""
    level_semester_raw: str = ""
    courses: List[ParsedCourse] = field(default_factory=list)
    # Calculated fields (filled after parsing)
    semester_gpa: Optional[float] = None
    total_hours: int = 0


@dataclass
class ParsedStudent:
    student_code: str
    name: str
    study_level_str: str = ""
    cumulative_percentage: Optional[float] = None
    enrollment_year: Optional[int] = None
    department: str = ""
    semesters: List[ParsedSemester] = field(default_factory=list)
    # Calculated after all semesters
    cumulative_gpa: Optional[float] = None
    total_passed_hours: int = 0


@dataclass
class ParsedData:
    students: List[ParsedStudent] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


# ─── Low-level cell helpers (KEPT from existing parser) ───────────────────────

def _clean(value) -> str:
    """Strip whitespace from cell value; return '' for None."""
    if value is None:
        return ""
    return str(value).strip()


def _get_cell(ws, col_letter: str, row_idx: int):
    """Read a single cell by column letter + row number."""
    try:
        return ws[f"{col_letter}{row_idx}"].value
    except Exception:
        return None


def _find_first_semester_row(ws) -> Tuple[Optional[int], Optional[int]]:
    """Scan first 100 rows to find the cell containing 'القسم/الشعبة'."""
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 80)):
            val = _clean(ws.cell(row, col).value)
            if "القسم/الشعبة" in val:
                return row, col
    return None, None


def _detect_offset(ws) -> Tuple[int, int]:
    """
    Returns (row_offset, col_offset) to correct for layout shifts between sheets.
    Sheet 1 may have different positioning than sheets 2+.
    """
    ref_row, ref_col = _find_first_semester_row(ws)
    if ref_row and ref_col:
        row_offset = ref_row - 40
        col_offset = ref_col - 11
        return row_offset, col_offset
    return 0, 0


def _get_cell_with_offset(ws, base_col_letter: str, base_row: int,
                           row_offset: int, col_offset: int):
    """Read cell with applied row/col offsets for layout correction."""
    base_col_num = openpyxl.utils.column_index_from_string(base_col_letter)
    new_col_num = base_col_num + col_offset
    if new_col_num < 1:
        return None
    new_col_letter = openpyxl.utils.get_column_letter(new_col_num)
    new_row = base_row + row_offset
    if new_row < 1:
        return None
    return _get_cell(ws, new_col_letter, new_row)


# ─── Excel File Parser ───────────────────────────────────────────────────────

class ExcelFileParser:
    """
    Wraps the existing column-based parser logic.
    Returns ParsedData with fully structured student objects.
    GPA is always CALCULATED from grades, never read from file.
    """

    # Column reference letters used by the original parser
    COLS = {
        "K": "K", "AH": "AH", "BO": "BO",
        "I": "I", "Y": "Y", "M": "M", "X": "X",
        "L": "L", "W": "W", "AN": "AN",
        "CD": "CD", "BC": "BC", "BU": "BU",
        "AL": "AL", "Z": "Z", "AC": "AC",
        "AS": "AS", "AU": "AU", "O": "O", "Q": "Q",
    }

    def __init__(self, grade_scale: Optional[Dict[str, float]] = None):
        """
        Args:
            grade_scale: {grade_letter: points} from Supabase grade_scale_items.
                         Falls back to FALLBACK_GRADE_POINTS if None.
        """
        self._grade_scale = grade_scale or FALLBACK_GRADE_POINTS

    @property
    def grade_scale(self) -> Dict[str, float]:
        return self._grade_scale

    @grade_scale.setter
    def grade_scale(self, value: Dict[str, float]) -> None:
        self._grade_scale = value

    # ── Public API ────────────────────────────────────────────────────────

    def parse_workbook(self, file_path: Path) -> ParsedData:
        """
        Opens workbook, processes each sheet as one student.
        Returns ParsedData with all students + any non-fatal errors.
        """
        result = ParsedData()
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except Exception as e:
            result.errors.append(f"فشل فتح ملف Excel: {e}")
            return result

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                student = self._parse_single_sheet(ws, sheet_name)
                if student is not None:
                    result.students.append(student)
            except Exception as e:
                msg = f"فشل تحليل ورقة '{sheet_name}': {e}"
                logger.error(msg, exc_info=True)
                result.errors.append(msg)

        wb.close()
        return result

    # ── Sheet-level parsing ───────────────────────────────────────────────

    def _parse_single_sheet(self, ws, sheet_name: str) -> Optional[ParsedStudent]:
        """
        Parse one sheet into a ParsedStudent.
        Returns None if sheet has no valid student data.
        """
        # 1. Read student header info
        student_code, name, study_level_str, cumulative_pct = self._parse_student_info(ws)
        if not student_code:
            logger.debug(f"ورقة '{sheet_name}' بدون كود طالب — تم التخطي")
            return None

        # 2. Parse semester blocks
        raw_semesters = self._parse_semesters(ws)
        if not raw_semesters:
            logger.warning(f"ورقة '{sheet_name}' (طالب {student_code}): لا توجد فصول دراسية")
            return None

        # 3. Build ParsedSemester objects with courses
        semesters: List[ParsedSemester] = []
        for idx, raw in enumerate(raw_semesters, start=1):
            level, term = self._parse_level_term(raw.get("level_semester", ""))
            courses = self._build_courses(raw.get("courses", []))

            sem = ParsedSemester(
                semester_number=idx,
                academic_year=raw.get("academic_year", ""),
                term=term,
                level=level,
                department=raw.get("department", ""),
                level_semester_raw=raw.get("level_semester", ""),
                courses=courses,
            )
            # Calculate semester GPA and total hours
            sem.semester_gpa = self._calculate_semester_gpa(courses, self._grade_scale)
            sem.total_hours = sum(
                c.credit_hours for c in courses
                if c.passed and c.credit_hours > 0
            )
            semesters.append(sem)

        # 4. Detect retakes across all semesters
        self._detect_retakes(semesters)

        # 5. Extract enrollment_year from first semester
        enrollment_year = self._extract_enrollment_year(semesters)

        # 6. Extract department from first semester
        department = semesters[0].department if semesters else ""

        # 7. Calculate cumulative GPA
        cumulative_gpa = self._calculate_cumulative_gpa(semesters)

        # 8. Calculate total passed hours
        total_passed_hours = sum(s.total_hours for s in semesters)

        # 9. Parse cumulative percentage
        cum_pct: Optional[float] = None
        if cumulative_pct:
            try:
                cum_pct = float(cumulative_pct)
            except (ValueError, TypeError):
                pass

        return ParsedStudent(
            student_code=student_code,
            name=name,
            study_level_str=study_level_str,
            cumulative_percentage=cum_pct,
            enrollment_year=enrollment_year,
            department=department,
            semesters=semesters,
            cumulative_gpa=cumulative_gpa,
            total_passed_hours=total_passed_hours,
        )

    # ── Student header parsing (KEPT from existing parser) ────────────────

    def _parse_student_info(self, ws) -> Tuple[str, str, str, str]:
        """
        Reads student code, name, study level, cumulative percentage
        from the first ~30 rows of the sheet.
        Returns: (student_code, name, study_level_str, cumulative_percentage_str)
        """
        student_code = ""
        name = ""
        study_level = ""
        cum_pct = ""

        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                v = _clean(cell.value)
                if v.startswith("كود الطالب :"):
                    student_code = v.replace("كود الطالب :", "").strip()
                elif v.startswith("أسم الطالب :"):
                    name = v.replace("أسم الطالب :", "").strip()
                elif v.startswith("مستوى الدراسة :"):
                    study_level = v.replace("مستوى الدراسة :", "").strip()
                elif v.startswith("النسبة(بحساب النقاط) :"):
                    cum_pct = v.replace("النسبة(بحساب النقاط) :", "").strip()

        return student_code, name, study_level, cum_pct

    # ── Semester block parsing (KEPT from existing parser) ────────────────

    def _parse_semesters(self, ws) -> List[Dict]:
        """
        Reads semester blocks using column offsets (K, AH, BO, etc.).
        Returns list of raw semester dicts with courses.
        """
        row_offset, col_offset = _detect_offset(ws)

        semesters = []
        current_semester = None
        current_courses = []
        cols = self.COLS

        start_row = max(1, 38 + row_offset)

        for row_idx in range(start_row, ws.max_row + 1):
            cell_k = _get_cell_with_offset(ws, cols["K"], row_idx, row_offset, col_offset)
            cell_ah = _get_cell_with_offset(ws, cols["AH"], row_idx, row_offset, col_offset)
            cell_bo = _get_cell_with_offset(ws, cols["BO"], row_idx, row_offset, col_offset)

            cell_k_clean = _clean(cell_k)

            # Semester header detection
            if "القسم/الشعبة :" in cell_k_clean:
                if current_semester is not None:
                    current_semester["courses"] = current_courses
                    semesters.append(current_semester)

                current_semester = {
                    "department": cell_k_clean.replace("القسم/الشعبة :", "").strip(),
                    "level_semester": _clean(cell_ah).replace("المستوى/الفصل :", "").strip(),
                    "academic_year": _clean(cell_bo).replace("العام الأكاديمي   :", "").replace("العام الأكاديمي:", "").strip(),
                }
                current_courses = []
                continue

            if current_semester is None:
                continue

            # Skip summary rows (تراكمى الفصل, نتيجة الفصل, etc.)
            # These are NOT used for GPA — we calculate from grades directly.
            cell_i_raw = _clean(_get_cell_with_offset(ws, cols["I"], row_idx, row_offset, col_offset))
            if cell_i_raw.startswith("الساعات المجتازة :") or cell_i_raw.startswith("تراكمى") or cell_i_raw.startswith("نتيجة"):
                continue

            # Skip header rows
            cell_cd = _clean(_get_cell_with_offset(ws, cols["CD"], row_idx, row_offset, col_offset))
            cell_bc = _clean(_get_cell_with_offset(ws, cols["BC"], row_idx, row_offset, col_offset))

            if cell_cd == "م" or cell_bc in ("المجمـــوع", "المجموع"):
                continue

            # Course row: cell_cd is a digit (row number) and cell_bc is the course name
            if cell_cd.isdigit() and cell_bc:
                course = {
                    "course_code": _clean(_get_cell_with_offset(ws, cols["BU"], row_idx, row_offset, col_offset)),
                    "course_name": cell_bc,
                    "passed_str": _clean(_get_cell_with_offset(ws, cols["I"], row_idx, row_offset, col_offset)),
                    "grade_letter": _clean(_get_cell_with_offset(ws, cols["O"], row_idx, row_offset, col_offset)),
                    "score": _clean(_get_cell_with_offset(ws, cols["Q"], row_idx, row_offset, col_offset)),
                    "hours": _clean(_get_cell_with_offset(ws, cols["AL"], row_idx, row_offset, col_offset)),
                }
                current_courses.append(course)

        # Don't forget the last semester
        if current_semester is not None:
            current_semester["courses"] = current_courses
            semesters.append(current_semester)

        return semesters

    # ── Course building ───────────────────────────────────────────────────

    def _build_courses(self, raw_courses: List[Dict]) -> List[ParsedCourse]:
        """Convert raw course dicts into ParsedCourse dataclasses."""
        courses: List[ParsedCourse] = []
        for c in raw_courses:
            grade_letter = c.get("grade_letter", "").strip()
            credit_hours = self._safe_int(c.get("hours", 0))
            score = self._safe_float(c.get("score"))
            passed = self._is_passed(grade_letter, c.get("passed_str", ""))
            grade_pts = self._grade_scale.get(grade_letter, FALLBACK_GRADE_POINTS.get(grade_letter, 0.0))

            courses.append(ParsedCourse(
                course_code=c.get("course_code", ""),
                course_name=c.get("course_name", ""),
                credit_hours=credit_hours,
                grade_letter=grade_letter,
                score=score,
                passed=passed,
                grade_points=grade_pts,
            ))
        return courses

    # ── GPA calculation (NEVER from file — always from grades) ────────────

    def _calculate_semester_gpa(
        self,
        courses: List[ParsedCourse],
        grade_scale: Dict[str, float],
    ) -> Optional[float]:
        """
        Calculate semester GPA from courses.
        Only includes courses with valid grade letters (not special symbols)
        and credit_hours > 0.
        Returns None if no gradeable courses exist.
        """
        total_points = 0.0
        total_hours = 0

        for c in courses:
            if c.grade_letter in SPECIAL_SYMBOLS:
                continue
            if c.credit_hours <= 0:
                continue
            pts = grade_scale.get(c.grade_letter, FALLBACK_GRADE_POINTS.get(c.grade_letter))
            if pts is None:
                continue
            total_points += pts * c.credit_hours
            total_hours += c.credit_hours

        if total_hours == 0:
            return None
        return round(total_points / total_hours, 4)

    def _calculate_cumulative_gpa(self, all_semesters: List[ParsedSemester]) -> Optional[float]:
        """Calculate cumulative GPA across all semesters from course-level data."""
        total_points = 0.0
        total_hours = 0

        for sem in all_semesters:
            for c in sem.courses:
                if c.grade_letter in SPECIAL_SYMBOLS:
                    continue
                if c.credit_hours <= 0:
                    continue
                pts = self._grade_scale.get(c.grade_letter, FALLBACK_GRADE_POINTS.get(c.grade_letter))
                if pts is None:
                    continue
                total_points += pts * c.credit_hours
                total_hours += c.credit_hours

        if total_hours == 0:
            return None
        return round(total_points / total_hours, 4)

    # ── Level/Term parsing ────────────────────────────────────────────────

    def _parse_level_term(self, level_semester_str: str) -> Tuple[Optional[int], Optional[str]]:
        """
        Parse "المستوى الاول/الفصل الدراسى الاول" → (1, "fall")
        """
        if not level_semester_str:
            return None, None

        level: Optional[int] = None
        term: Optional[str] = None

        parts = level_semester_str.split("/")

        # Level: usually first part
        level_part = parts[0].strip() if len(parts) >= 1 else ""
        for keyword, val in LEVEL_MAP.items():
            if keyword in level_part:
                level = val
                break

        # Term: usually second part onwards
        term_part = "/".join(parts[1:]).strip() if len(parts) >= 2 else ""
        # Try exact match first
        for keyword, val in TERM_MAP.items():
            if keyword in term_part:
                term = val
                break

        # Fallback: check full string for term keywords
        if term is None:
            for keyword, val in TERM_MAP.items():
                if keyword in level_semester_str:
                    term = val
                    break

        return level, term

    # ── Enrollment year extraction ────────────────────────────────────────

    def _extract_enrollment_year(self, semesters: List[ParsedSemester]) -> Optional[int]:
        """
        Enrollment year = SECOND number from first semester's academic_year.
        E.g. "2020-2021" → 2021
        """
        if not semesters:
            return None
        ay = semesters[0].academic_year
        if not ay:
            return None
        # Try YYYY-YYYY pattern
        match = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", ay)
        if match:
            try:
                return int(match.group(2))
            except (ValueError, IndexError):
                pass
        # Fallback: just extract any 4-digit year
        match = re.search(r"(\d{4})", ay)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                pass
        return None

    # ── Retake detection ──────────────────────────────────────────────────

    @staticmethod
    def _detect_retakes(all_semesters: List[ParsedSemester]) -> None:
        """
        Mutates courses in place, setting is_retake and retake_count.
        Process semesters in order (semester_number ascending).
        """
        seen_codes: Dict[str, int] = {}
        for semester in sorted(all_semesters, key=lambda s: s.semester_number):
            for course in semester.courses:
                code = course.course_code
                if code and code not in ("", "N/A"):
                    count = seen_codes.get(code, 0)
                    course.is_retake = count > 0
                    course.retake_count = count
                    seen_codes[code] = count + 1

    # ── Helpers ───────────────────────────────────────────────────────────

    @staticmethod
    def _is_passed(grade_letter: str, passed_str: str = "") -> bool:
        """Determine if a course is passed based on grade letter."""
        if grade_letter in SPECIAL_SYMBOLS:
            return False
        if grade_letter == "F":
            return False
        if not grade_letter:
            return passed_str == "نعم"
        # Any real grade with points > 0 is passed
        pts = FALLBACK_GRADE_POINTS.get(grade_letter)
        if pts is not None:
            return pts > 0
        # Unknown grade — check Arabic passed_str
        return passed_str == "نعم"

    @staticmethod
    def _safe_int(value, default: int = 0) -> int:
        try:
            return int(float(str(value).strip() or "0"))
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _safe_float(value) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(str(value).strip())
        except (ValueError, TypeError):
            return None


# Alias for backward compatibility and package initializer compatibility
FileParser = ExcelFileParser

